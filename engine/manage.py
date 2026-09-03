#!/usr/bin/env python3
"""Правка самого чек-листа: добавить, убрать, изменить вопрос; импорт из xlsx-шаблона IMF.

  manage.py where                          показать, какая копия чек-листа сейчас рабочая
  manage.py fork [--dir checklist_data]    сделать локальную копию, которую можно править
  manage.py list [--process ...]           все вопросы, включая выключенные
  manage.py add --id PRD42 --process "Работа с продуктом" --question-ru "..." \
                [--question-en "..."] --levels D1;D2 --zones fridge,freezer --days 10 \
                [--criteria "D1: ...\\nD2: ..."]
  manage.py remove PRD42 [--hard]          выключить вопрос (или удалить строку совсем)
  manage.py restore PRD42                  включить обратно
  manage.py edit PRD42 --levels D1;D2;D3 --days 7 --zones "*"
  manage.py zone-add --code terrace --name-ru "Терраса" --name-en "Terrace"
  manage.py zone-remove terrace
  manage.py validate                       проверить целостность файлов
  manage.py import-xlsx файл.xlsx [--keep-zones] [--drop-extra-columns]
        пересобрать чек-лист из выгрузки шаблона IMF (Template_CL). --keep-zones
        сохраняет уже расставленные зоны для вопросов с совпадающей формулировкой.
        --drop-extra-columns — согласие потерять колонки, которых нет в шаблоне.

Выключенный вопрос (kind=off) остаётся в файле, но не предлагается при проверке и не
участвует в расчёте — так историю правок видно, а вернуть пункт можно одной командой.

Колонки, которых движок не знает, при правке методики сохраняются: чек-лист и зоны —
данные управляющей компании, и `manage.py` не вправе выбрасывать из них то, чего не
читает сам. Исключение одно — `import-xlsx`, см. выше.
"""
import argparse, csv, json, os, re, shutil, sys

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)
from audit import (  # noqa: E402
    ZONE_SHARE_TOLERANCE, ZONE_SHARE_TOTAL, active_dir, data_path, load_zones,
)

PLUGIN_DATA = os.path.normpath(os.path.join(HERE, os.pardir, "data"))
FIELDS = ["id", "kind", "process_ru", "process_en", "question_ru", "question_en", "levels", "zones", "days"]
ZONE_FIELDS = ["code", "name_ru", "name_en", "share_pct"]
# D0 — не класс нарушения, а приём: информационная запись живёт среди findings
# с нулевым вычетом (docs/02-domain.md). Уровня не знала только эта проверка, и
# три боевых пункта (INF09-INF11) делали `validate` красным всегда. Проверка,
# которая кричит на исправных данных, так же бесполезна, как та, что не может
# упасть: её перестают запускать — вместе со всем, что она ловит.
VIOLATION_LEVELS = {"D1", "D2", "D3"}
LEVELS = VIOLATION_LEVELS | {"D0"}


def target_dir(create=False, name="checklist_data"):
    d = active_dir()
    if os.access(d, os.W_OK) and d != PLUGIN_DATA:
        return d
    if os.access(PLUGIN_DATA, os.W_OK) and not create:
        return PLUGIN_DATA
    dst = os.path.join(os.getcwd(), name)
    if not os.path.isdir(dst):
        os.makedirs(dst, exist_ok=True)
        for f in ("checklist.csv", "criteria.md", "zones.csv", "scoring.json"):
            src = data_path(f)
            if os.path.exists(src):
                shutil.copy2(src, os.path.join(dst, f))
        print(f"Создана рабочая копия чек-листа: {dst}", file=sys.stderr)
    return dst


def read_rows(d=None):
    p = os.path.join(d, "checklist.csv") if d else data_path("checklist.csv")
    with open(p, encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def foreign_columns(path, known, rows=()):
    """Колонки файла методики, которых движок не знает, — в порядке появления.

    Всё, что не перечислено в `known`, — данные управляющей компании: движок их
    не читает, но и распоряжаться ими не вправе. Смотрим и в заголовок файла на
    диске, и в сами строки: заголовок нужен, когда строк не осталось (файл с
    одной шапкой), строки — когда пишем не в тот файл, из которого читали.

    Ключ `None` пропускается намеренно: так `csv.DictReader` складывает хвост
    строки, в которой значений больше, чем колонок в шапке. Имени у такого
    хвоста нет, колонкой он не является, и записать его некуда.
    """
    seen = []

    def подобрать(имена):
        for k in имена:
            if k is not None and k not in known and k not in seen:
                seen.append(k)

    if os.path.exists(path):
        with open(path, encoding="utf-8-sig") as f:
            подобрать(next(csv.reader(f), []))
    for r in rows:
        подобрать(r)
    return seen


def write_csv_rows(path, rows, known, keep_foreign=True):
    """Перезаписать файл методики, не потеряв чужие колонки.

    Раньше здесь стоял фиксированный список колонок, и любая колонка, заведённая
    управляющей компанией сверх него, исчезала при первой же правке методики —
    молча и безвозвратно (методика лежит вне git, D002). Настолько не теория,
    что порядок обхода точки (T061) пришлось делать отдельным файлом
    `data/route.csv`: колонкой в чек-листе он не пережил бы первую же правку.

    Известные движку колонки идут первыми и в прежнем порядке — формат файла от
    сохранения чужого не перетасовывается. `keep_foreign=False` оставляет только
    известные: это `import-xlsx`, где чек-лист пересобирается из чужой выгрузки
    целиком и значений для чужих колонок взять неоткуда (задача T109).
    """
    fields = [*known, *(foreign_columns(path, known, rows) if keep_foreign else [])]
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        for r in rows:
            w.writerow({k: r.get(k, "") for k in fields})


def write_rows(rows, d, keep_foreign=True):
    write_csv_rows(os.path.join(d, "checklist.csv"), rows, FIELDS, keep_foreign=keep_foreign)


def set_criteria(d, qid, text):
    p = os.path.join(d, "criteria.md")
    body = open(p, encoding="utf-8").read() if os.path.exists(p) else "# Критерии нарушений по вопросам\n"
    body = re.sub(rf"\n## {re.escape(qid)}\n.*?(?=\n## |\Z)", "", body, flags=re.S)
    body = body.rstrip() + f"\n\n## {qid}\n{text.strip()}\n"
    open(p, "w", encoding="utf-8").write(body)


def cmd_where(a):
    d = active_dir()
    print(f"рабочая копия: {d}")
    print(f"файлы плагина: {PLUGIN_DATA}")
    print(f"доступна на запись: {'да' if os.access(d, os.W_OK) else 'нет'}")
    rows = read_rows()
    from collections import Counter
    print("вопросов:", len(rows), dict(Counter(r.get("kind", "") for r in rows)))


def cmd_fork(a):
    d = target_dir(create=True, name=a.dir or "checklist_data")
    print(d)


def cmd_list(a):
    for r in read_rows():
        if a.process and a.process.lower() not in r["process_ru"].lower():
            continue
        if a.q and a.q.lower() not in (r["question_ru"] + r.get("question_en", "")).lower():
            continue
        flag = "" if r.get("kind") == "violation" else f" [{r.get('kind')}]"
        print(f"{r['id']}{flag} | {r['process_ru']} | {r['question_ru'][:95]} | {r.get('levels','')} | {r.get('zones','')} | {r.get('days','')}д")


def cmd_add(a):
    d = target_dir(create=True)
    rows = read_rows(d if os.path.exists(os.path.join(d, "checklist.csv")) else None)
    ids = {r["id"] for r in rows}
    qid = (a.id or "").strip().upper()
    if not qid:
        pre = re.sub(r"[^A-ZА-Я]", "", (a.process or "NEW").upper())[:3] or "NEW"
        n = 1
        while f"{pre}{n:02d}" in ids:
            n += 1
        qid = f"{pre}{n:02d}"
    if qid in ids:
        sys.exit(f"{qid} уже есть. Выберите другой id или используйте edit.")
    zc = {z["code"] for z in load_zones()}
    zones = (a.zones or "*").strip()
    if zones != "*":
        bad = [z for z in zones.split(",") if z.strip() and z.strip() not in zc]
        if bad:
            sys.exit(f"Неизвестные зоны: {', '.join(bad)}. Доступны: {', '.join(sorted(zc))}")
    levels = ";".join(x.strip().upper() for x in re.split(r"[;,]", a.levels or "D1") if x.strip())
    row = {"id": qid, "kind": a.kind or "violation", "process_ru": a.process or "",
           "process_en": a.process_en or a.process or "", "question_ru": a.question_ru or "",
           "question_en": a.question_en or "", "levels": levels, "zones": zones,
           "days": a.days if a.days is not None else 10}
    rows.append(row)
    write_rows(rows, d)
    if a.criteria:
        set_criteria(d, qid, a.criteria)
    print(f"добавлен {qid} → {os.path.join(d, 'checklist.csv')}")


def cmd_remove(a):
    d = target_dir(create=True)
    rows = read_rows()
    hit = [r for r in rows if r["id"].upper() == a.id.upper()]
    if not hit:
        sys.exit(f"Нет вопроса {a.id}")
    if a.hard:
        rows = [r for r in rows if r["id"].upper() != a.id.upper()]
    else:
        for r in hit:
            r["kind"] = "off"
    write_rows(rows, d)
    print(("удалён " if a.hard else "выключен ") + a.id.upper())


def cmd_restore(a):
    d = target_dir(create=True)
    rows = read_rows()
    for r in rows:
        if r["id"].upper() == a.id.upper():
            r["kind"] = "violation"
    write_rows(rows, d)
    print("включён " + a.id.upper())


def cmd_edit(a):
    d = target_dir(create=True)
    rows = read_rows()
    found = False
    for r in rows:
        if r["id"].upper() != a.id.upper():
            continue
        found = True
        for f, v in (("process_ru", a.process), ("process_en", a.process_en),
                     ("question_ru", a.question_ru), ("question_en", a.question_en),
                     ("zones", a.zones), ("kind", a.kind)):
            if v is not None:
                r[f] = v
        if a.levels is not None:
            r["levels"] = ";".join(x.strip().upper() for x in re.split(r"[;,]", a.levels) if x.strip())
        if a.days is not None:
            r["days"] = a.days
    if not found:
        sys.exit(f"Нет вопроса {a.id}")
    write_rows(rows, d)
    if a.criteria:
        set_criteria(d, a.id.upper(), a.criteria)
    print("обновлён " + a.id.upper())


def cmd_zone_add(a):
    d = target_dir(create=True)
    p = os.path.join(d, "zones.csv")
    rows = list(csv.DictReader(open(p, encoding="utf-8-sig"))) if os.path.exists(p) else \
        [dict(r, share_pct=r["share_pct"]) for r in load_zones()]
    if any(r["code"] == a.code for r in rows):
        sys.exit("такая зона уже есть")
    rows.append({"code": a.code, "name_ru": a.name_ru, "name_en": a.name_en or a.name_ru, "share_pct": 0})
    share = round(100 / len(rows), 4)
    for r in rows:
        r["share_pct"] = share
    write_csv_rows(p, rows, ZONE_FIELDS)
    print(f"зон стало {len(rows)}, доля каждой {share:g}%")


def cmd_zone_remove(a):
    d = target_dir(create=True)
    p = os.path.join(d, "zones.csv")
    rows = [r for r in csv.DictReader(open(data_path("zones.csv"), encoding="utf-8-sig")) if r["code"] != a.code]
    share = round(100 / len(rows), 4)
    for r in rows:
        r["share_pct"] = share
    write_csv_rows(p, rows, ZONE_FIELDS)
    cl = read_rows()
    touched = []
    for r in cl:
        zs = [z for z in (r.get("zones") or "").split(",") if z.strip() and z.strip() != a.code]
        if (r.get("zones") or "").find(a.code) >= 0 and (r.get("zones") or "") != "*":
            r["zones"] = ",".join(zs) or "*"
            touched.append(r["id"])
    write_rows(cl, d)
    print(f"зона {a.code} убрана, доля каждой из {len(rows)} — {share:g}%")
    if touched:
        print("вопросы, у которых зона убрана из списка:", ", ".join(touched))


def levels_of(row):
    """Классы пункта из колонки `levels` — множеством, в верхнем регистре."""
    return {x.strip().upper() for x in re.split(r"[;,]", row.get("levels", "")) if x.strip()}


def zone_rows_raw():
    """Зоны как они лежат в файле, без проверок движка.

    `load_zones()` здесь звать нельзя. С T103 он сам завершает процесс на
    несходящихся долях, а с T106 — ещё и на нечисловых: для расчёта оценки это
    правильно, но `validate` от такого умирает раньше первой напечатанной
    строки. Инструмент, которым человек чинит методику, обязан показать ВСЕ
    проблемы разом, а не падать на первой.
    """
    path = data_path("zones.csv")
    with open(path, encoding="utf-8-sig") as f:
        return [r for r in csv.DictReader(f) if (r.get("code") or "").strip()]


def zone_problems(zrows):
    """Что не так с долями зон. Допуск — тот же, что у движка, и это важно.

    Раньше здесь стояло `abs(zs - 100) > 0.05` против 0.005 в движке, и сумма
    сходилась только потому, что считалась по УЖЕ нормализованным долям:
    `load_zones()` переписывал их на равные, и проверка всегда получала ровно
    100. Упасть она не могла ни на каких данных. Разойдись допуски теперь —
    человек починит методику по зелёному `validate`, а `score` её не посчитает.
    """
    problems = []
    total = 0.0
    readable = True
    for r in zrows:
        code = (r.get("code") or "").strip()
        raw = (r.get("share_pct") or "").strip()
        if not raw:
            problems.append(f"зона {code}: доля не заполнена — сумму долей проверить не на чем")
            readable = False
            continue
        try:
            total += float(raw)
        except ValueError:
            problems.append(f"зона {code}: доля «{raw}» не число — сумму долей проверить не на чем")
            readable = False
    if zrows and readable and abs(total - ZONE_SHARE_TOTAL) > ZONE_SHARE_TOLERANCE:
        problems.append(
            f"доли зон в сумме дают {total:g}%, а не {ZONE_SHARE_TOTAL:g}% "
            f"(движок такую методику считать откажется)")
    return problems


def cmd_validate(a):
    problems = []
    rows = read_rows()
    zrows = zone_rows_raw()
    zc = {(r.get("code") or "").strip() for r in zrows}
    seen = set()
    for r in rows:
        qid = r.get("id", "").strip()
        if not qid:
            problems.append("строка без id")
            continue
        if qid in seen:
            problems.append(f"{qid}: дубль id")
        seen.add(qid)
        if r.get("kind") not in ("violation", "info", "aggregate", "off"):
            problems.append(f"{qid}: неизвестный kind «{r.get('kind')}»")
        if r.get("kind") == "violation":
            lv = levels_of(r)
            if not lv:
                problems.append(f"{qid}: не указан ни один уровень D")
            for x in lv:
                if x not in LEVELS:
                    problems.append(f"{qid}: неизвестный уровень «{x}»")
            z = (r.get("zones") or "").strip()
            if z and z != "*":
                for x in z.split(","):
                    if x.strip() and x.strip() not in zc:
                        problems.append(f"{qid}: неизвестная зона «{x.strip()}»")
            if not r.get("question_ru"):
                problems.append(f"{qid}: пустая формулировка")
            try:
                int(float(r.get("days") or 0))
            except ValueError:
                problems.append(f"{qid}: срок не число")
    crit = open(data_path("criteria.md"), encoding="utf-8").read()
    have = set(re.findall(r"\n## (\S+)\n", crit))
    for r in rows:
        if r.get("kind") == "violation" and levels_of(r) & VIOLATION_LEVELS and r["id"] not in have:
            problems.append(f"{r['id']}: нет критериев в criteria.md (совет: добавьте, иначе распознавание по фото будет угадывать)")
    problems.extend(zone_problems(zrows))
    print(f"вопросов: {len(rows)}, зон: {len(zc)}")
    if problems:
        print("\nПроблемы:")
        for p in problems:
            print(" -", p)
        sys.exit(1)
    print("Всё в порядке.")


def cmd_import(a):
    try:
        import openpyxl
    except ImportError:
        sys.exit("Нужен openpyxl: pip install openpyxl --break-system-packages")
    d = target_dir(create=True)
    # Единственная команда, которая пересобирает чек-лист целиком из чужой
    # выгрузки: значений для колонок управляющей компании в шаблоне IMF нет и
    # взяться им неоткуда. Сохранить их здесь нельзя — можно только не потерять
    # молча, поэтому отказ с именами колонок и явный флаг для того, кто согласен
    # их потерять (T109).
    cl_path = os.path.join(d, "checklist.csv")
    foreign = foreign_columns(cl_path, FIELDS)
    if foreign and not a.drop_extra_columns:
        sys.exit(
            f"В чек-листе есть колонки, которых нет в шаблоне IMF: {', '.join(foreign)}. "
            f"Файл: {cl_path}. Импорт пересобирает чек-лист целиком из выгрузки, значений "
            f"для этих колонок в ней нет, а методика лежит вне git — восстановить их будет "
            f"неоткуда. Сохраните копию файла, затем повторите с --drop-extra-columns."
        )
    old = {}
    if a.keep_zones:
        for r in read_rows():
            old[re.sub(r"\s+", " ", r["question_ru"]).strip().lower()] = r
    wb = openpyxl.load_workbook(a.path, data_only=True)
    ws = wb["Template_CL"] if "Template_CL" in wb.sheetnames else wb.worksheets[0]
    rows_all = list(ws.iter_rows(values_only=True))
    hdr_i = None
    for i, r in enumerate(rows_all[:8]):
        cells = [str(c or "").strip().lower() for c in r]
        if any(c.startswith("question") or c.startswith("вопрос") for c in cells):
            hdr_i = i
            break
    if hdr_i is None:
        sys.exit("Не нашёл строку заголовков (в ней должна быть колонка Question / Вопрос)")
    hdr = [re.sub(r"\s+", " ", str(c or "")).strip().lower() for c in rows_all[hdr_i]]

    def col(*names, en=False):
        for i, h in enumerate(hdr):
            if not h:
                continue
            is_en = "(en)" in h
            if en != is_en:
                continue
            for n in names:
                if h.startswith(n):
                    return i
        return None

    ci = {"proc_ru": col("process", "процесс"), "q_ru": col("question", "вопрос"),
          "proc_en": col("process", en=True), "q_en": col("question", en=True),
          "days": col("days for fixing", "дней"), "hint": col("hint", "подсказка"),
          "ans": col("custom answer", "свой вариант", "вариант ответа")}
    for k in ("proc_ru", "q_ru", "ans"):
        if ci[k] is None:
            sys.exit(f"В файле не нашлась обязательная колонка ({k}). Проверьте, что это шаблон Template_CL.")
    ci["proc_en"] = ci["proc_en"] if ci["proc_en"] is not None else ci["proc_ru"]
    ci["q_en"] = ci["q_en"] if ci["q_en"] is not None else ci["q_ru"]
    raw = [r for r in rows_all[hdr_i + 1:] if r and len(r) > ci["q_ru"] and str(r[ci["q_ru"]] or "").strip()]
    if not raw:
        sys.exit("В файле нет строк с вопросами")

    agg = {}
    for r in raw:
        key = (str(r[ci["proc_ru"]]).strip(), re.sub(r"\s+", " ", str(r[ci["q_ru"]])).strip())
        e = agg.setdefault(key, {"proc_en": str(r[ci["proc_en"]] or "").strip(),
                                 "q_en": re.sub(r"\s+", " ", str(r[ci["q_en"]] or "")).strip(),
                                 "days": r[ci["days"]], "hint": str(r[ci["hint"]] or "").strip(), "ans": []})
        e["ans"].append(str(r[ci["ans"]] or "").strip())
    TR = {"а":"a","б":"b","в":"v","г":"g","д":"d","е":"e","ё":"e","ж":"z","з":"z","и":"i","й":"i",
          "к":"k","л":"l","м":"m","н":"n","о":"o","п":"p","р":"r","с":"s","т":"t","у":"u","ф":"f",
          "х":"h","ц":"c","ч":"c","ш":"s","щ":"s","ы":"y","э":"e","ю":"u","я":"a","ъ":"","ь":""}

    def ascii_prefix(name):
        words = re.findall(r"[A-Za-zА-Яа-яЁё]+", name)
        letters = "".join(TR.get(w[0].lower(), w[0].lower()) for w in words)
        letters = re.sub(r"[^a-z]", "", letters)
        if not letters:
            letters = "gen"
        return (letters * 3)[:3].upper()

    pre_by_proc, rows, crit = {}, [], []
    used = set()
    for (proc, q), e in agg.items():
        pre = pre_by_proc.get(proc)
        if not pre:
            base = ascii_prefix(proc)
            pre = base
            i = 1
            while pre in pre_by_proc.values():
                pre = base[:2] + str(i); i += 1
            if len(pre) < 2:
                pre = (pre + "XX")[:3]
            pre_by_proc[proc] = pre
        n = 1
        while f"{pre}{n:02d}" in used:
            n += 1
        qid = f"{pre}{n:02d}"; used.add(qid)
        levels = [dd for dd in ("D1", "D2", "D3") if any(x.replace("\n", " ").strip() == f"Нет {dd}" for x in e["ans"])]
        if q.startswith("Зафиксировано критическое нарушение D3 в процессе"):
            kind = "aggregate"
        elif proc.rstrip(".").lower() in ("информация", "information"):
            kind = "info"
        else:
            kind = "violation"
            levels = levels or ["D1"]
        prev = old.get(q.lower())
        zones = prev.get("zones") if prev else "*"
        rows.append({"id": qid, "kind": kind, "process_ru": proc.rstrip("."), "process_en": e["proc_en"],
                     "question_ru": q, "question_en": e["q_en"], "levels": ";".join(levels),
                     "zones": zones if kind == "violation" else "",
                     "days": int(float(e["days"])) if str(e["days"] or "").replace(".", "").isdigit() else 10})
        if e["hint"] and kind == "violation":
            crit.append((qid, e["hint"].strip().strip('"')))
    write_rows(rows, d, keep_foreign=False)
    with open(os.path.join(d, "criteria.md"), "w", encoding="utf-8") as f:
        f.write("# Критерии нарушений по вопросам\n")
        for qid, h in crit:
            f.write(f"\n## {qid}\n{h}\n")
    noz = [r["id"] for r in rows if r["kind"] == "violation" and r["zones"] == "*"]
    print(f"импортировано вопросов: {len(rows)} → {d}")
    if noz:
        print(f"без явных зон (доступны все, при фиксации зону нужно уточнять): {len(noz)}")
        print("  " + ", ".join(noz[:25]) + (" …" if len(noz) > 25 else ""))


def main():
    p = argparse.ArgumentParser()
    s = p.add_subparsers(dest="cmd", required=True)
    s.add_parser("where").set_defaults(fn=cmd_where)
    f = s.add_parser("fork"); f.add_argument("--dir"); f.set_defaults(fn=cmd_fork)
    l = s.add_parser("list"); l.add_argument("--process"); l.add_argument("--q"); l.set_defaults(fn=cmd_list)
    ad = s.add_parser("add")
    for x in ("id", "process", "process-en", "question-ru", "question-en", "levels", "zones", "criteria", "kind"):
        ad.add_argument("--" + x)
    ad.add_argument("--days", type=int); ad.set_defaults(fn=cmd_add)
    rm = s.add_parser("remove"); rm.add_argument("id"); rm.add_argument("--hard", action="store_true"); rm.set_defaults(fn=cmd_remove)
    rs = s.add_parser("restore"); rs.add_argument("id"); rs.set_defaults(fn=cmd_restore)
    ed = s.add_parser("edit"); ed.add_argument("id")
    for x in ("process", "process-en", "question-ru", "question-en", "levels", "zones", "criteria", "kind"):
        ed.add_argument("--" + x)
    ed.add_argument("--days", type=int); ed.set_defaults(fn=cmd_edit)
    za = s.add_parser("zone-add"); za.add_argument("--code", required=True); za.add_argument("--name-ru", required=True); za.add_argument("--name-en"); za.set_defaults(fn=cmd_zone_add)
    zr = s.add_parser("zone-remove"); zr.add_argument("code"); zr.set_defaults(fn=cmd_zone_remove)
    s.add_parser("validate").set_defaults(fn=cmd_validate)
    im = s.add_parser("import-xlsx"); im.add_argument("path"); im.add_argument("--keep-zones", action="store_true")
    im.add_argument("--drop-extra-columns", action="store_true"); im.set_defaults(fn=cmd_import)
    a = p.parse_args()
    a.fn(a)


if __name__ == "__main__":
    main()
